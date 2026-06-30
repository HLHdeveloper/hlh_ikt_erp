# -*- coding: utf-8 -*-
# Import de aulas (gela_teoria_ids + tailerra_ids) en op.subject desde
# "Copia de Ordutegiak sortzeko.xlsx". Idempotente.
#   docker exec -i odoo19 odoo shell -d kudeaketa --no-http < import_aulak.py
import openpyxl, re

APPLY = True  # aplicado 2026-06-30
XLSX = '/tmp/ordutegiak.xlsx'
# orden de prioridad: KOG al FINAL para que sobrescriba a ELEK (decisión usuario)
SHEETS = ['ELEK','MEK','INF','AST','ZIAORIEN','ING','FOL','KOG']

# ---------- 1) aulas nuevas ----------
NEW_AULAS = [('1-L04','1-L04','gela'), ('1-P01','1-P01','gela'), ('1-P02','1-P02','gela')]
Classroom = env['op.classroom']
for code, name, mota in NEW_AULAS:
    if not Classroom.search([('code','=',code)], limit=1):
        print("  [crear aula] %s (%s)" % (code, mota))
        if APPLY:
            Classroom.create({'code':code,'name':name,'gela_mota':mota,'irakasgela':True})

cl = Classroom.search([('irakasgela','=',True)])
by_code = {c.code.upper(): c for c in cl}
by_name = {}
for c in cl:
    by_name.setdefault((c.name or '').strip().upper(), c)
AULA_OVR = {'1-L02C':'1-L02','1-L02B':'1-L02-B','1-L02A':'1-L02-A',
            '1L02C':'1-L02','1L02B':'1-L02-B','1L02A':'1-L02-A',
            '1-R01D':'1-R01-D','3-L03/A':'3-L03-A',
            'EXTRA':'2-L09','INFOR2':'4-L02','INFOR3':'4-L03'}
def aula_norm(t): return re.sub(r'^(\d)([A-Z])', r'\1-\2', t.strip().upper())
def aula(tok):
    if tok in AULA_OVR: return by_code.get(AULA_OVR[tok].upper())
    u = tok.strip().upper()
    return by_code.get(u) or by_code.get(aula_norm(tok)) or by_name.get(u)
def aula_cells(cells):
    out = []
    for cell in cells:
        for tok in re.split(r'\s+/\s+', str(cell).strip()):
            tok = tok.strip()
            if not tok: continue
            c = aula(tok)
            if c and c.id not in [x.id for x in out]: out.append(c)
    return out

# ---------- 2) alias de código de módulo (del banaketa) ----------
Subject = env['op.subject']
subj_codes = set(s.code for s in Subject.search([]))

def fix_taldea(tal):
    return tal.replace('OLHELE1','OLHELE3')

# alias del SUFIJO (módulo). Algunos dependen del taldea.
def alias_suffix(tal, name):
    n = name.strip()
    y = tal[0] if (tal and tal[0].isdigit()) else ''
    # contexto OLHELE (eleanitza): sufijos por año, IMRTD/BEZ/EEE literales
    if 'OLHELE' in tal:
        m = {'DIGI':'DIGI_'+y, 'PROSOL':'PROSOL_'+y, 'IED':'IED_'+y,
             'IED1':'IED_1', 'IED2':'IED_2', 'PROIEK':'PRO',
             'BTIETKMM':'TENBA', 'ALMAC':'ALMA', 'FUNDE':'FE'}
        if n in m: return m[n]
        if n.startswith('TUTO'): return 'TUTO_'+y
        return n
    SEA = 'SEA' in tal
    table = {'PSAD2':'PSAD_2','HAUT1':'HAUT_1','HAUT2':'HAUT_2','EETAK':'ICTVE',
             'IA':'HAUT_1','PLC':'HAUT_2','EEEL':'EEE','IMRTD':'IMRT','BTIETKMM':'TENBA',
             'FUNDE':'FE','ALMAC':'ALMA','ING_A':'ING_P_2','MOD_OPT_2':'HAUT_2',
             'ID':'IDOM','PROG':'HAUT_2'}
    # REGLA: las optativas se mantienen SIEMPRE como HAUT_1/HAUT_2 en Odoo (el
    # nombre real puede cambiar cada curso). El Excel trae el nombre real de la
    # optativa -> se mapea aquí a HAUT_x (IA/PLC/MOD_OPT_2/PROG...). NO renombrar
    # el módulo en Odoo; solo añadir el nombre nuevo a este alias cada curso.
    if SEA and n == 'DIGI': return 'PSAD_2'
    if n.startswith('TUTO'): return 'TUTO_'+y
    if n == 'EAE' and tal == '1MLE2': return 'EAEL'
    if n in table: return table[n]
    # EIP romano/arabe
    if n in ('EIP_1','EIP_2'):
        arabe_tal = tal in ('1MSS2','2MSS2','1SEA3','2SEA3')
        if tal == '1IEA2D' and n == 'EIP_1': return 'IPE_I'
        if arabe_tal: return n
        return {'EIP_1':'EIP_I','EIP_2':'EIP_II'}[n]
    if n == 'IPE_1': return 'IPE_I'
    return n

def build_code(tal, raw):
    raw = re.sub(r'\s*\(.*?\)\s*', '', str(raw)).strip()  # quitar anotaciones " (EEE)"
    tal = fix_taldea(tal)
    # módulos FG_* : código literal sin prefijo de taldea
    if raw.upper().startswith('FG_'):
        return raw
    # ¿código completo? (empieza por dígito y tiene '_')
    if re.match(r'^\d', raw) and '_' in raw:
        m = re.match(r'^(\d[A-Z0-9]+?)_(.+)$', raw)
        if m:
            t2 = fix_taldea(m.group(1)); suf = alias_suffix(t2, m.group(2))
            cand = '%s_%s' % (t2, suf)
            if cand in subj_codes: return cand
            if raw in subj_codes: return raw
            return cand
        return raw
    # código pelado -> taldea + sufijo
    suf = alias_suffix(tal, raw)
    return '%s_%s' % (tal, suf)

# ---------- 3) parsear xlsx por bloques ----------
def clean_aula(v):
    if v is None: return None
    s = str(v).strip()
    return None if (s=='' or s=='?') else s

assign = {}   # code -> (teoria_ids, prac_ids)  (último gana)
notfound = []
wb = openpyxl.load_workbook(XLSX, data_only=True)
for sh in SHEETS:
    ws = wb[sh]
    taldea = None
    teoria_cols = [11, 12, 13, 14]   # por defecto; se recalcula en cada cabecera
    prac_col = 15
    for r in range(1, ws.max_row+1):
        a = ws.cell(row=r, column=1).value
        if a is None: continue
        a = str(a).strip()
        if '—' in a:                       # cabecera de bloque
            taldea = a.split('—')[0].strip().split('+')[0].strip()
            continue
        if a == 'Módulo':                  # fila de cabecera: detectar columnas de aula
            tc, pc = [], None
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str):
                    vs = v.strip()
                    if vs.startswith('Aula teoria'): tc.append(c)
                    elif vs == 'Aula práctica': pc = c
            if tc: teoria_cols = tc
            if pc: prac_col = pc
            continue
        if a == '' or a.upper().startswith('TOTAL') or a.startswith('Planificación') or a.startswith('Horas de clase'):
            continue
        b = ws.cell(row=r, column=2).value
        if not isinstance(b,(int,float)): continue
        teoria = [clean_aula(ws.cell(row=r,column=c).value) for c in teoria_cols]
        prac = [clean_aula(ws.cell(row=r,column=prac_col).value)]
        teoria = [x for x in teoria if x]; prac = [x for x in prac if x]
        code = build_code(taldea or '', a)
        if code not in subj_codes:
            notfound.append("%s|%s|%s->%s" % (sh, taldea, a, code))
            continue
        tro = [c.id for c in aula_cells(teoria)]
        pra = [c.id for c in aula_cells(prac)]
        assign[code] = (tro, pra)   # KOG (último) sobrescribe

# ---------- correcciones manuales (revisión con usuario 2026-06-30) ----------
# Ganan sobre lo leído del Excel (se aplican al final).
# 1ELE1_EEE: usar el dato de ELEK (3-L01), no el de KOG (2-L09).
# 2MLE2_HAUT_2 / 2MLE2_MUMA: sin aula en Excel -> teoría 1-L02 + 1-L02-B.
FORCE = {
    '1ELE1_EEE': (['3-L01'], []),
    # 2MLE2_HAUT_2 no está en el Excel -> aula dada por el usuario.
    '2MLE2_HAUT_2': (['1-L02', '1-L02-B'], []),
    # INF: el bloque "INF" del Excel está vacío -> aulas dadas por el usuario.
    '1INF4_MUNTAIA': (['4-L01'], []),
    '1INF4_SAREAK': (['4-L01'], []),
    '1INF4_KONF': (['4-L01'], []),
    '1INF4_TUTO_1': (['4-L01'], []),
    '2INF4_SISTEMAK': (['2-L09', '2-L07'], []),
    '2INF4_TUTO_2': (['2-L09', '2-L07'], []),
}
# 3OLHELE3: grupo NO impartido en 26-27 -> sin aulas (omitir).
SKIP_PREFIX = ('3OLHELE3_',)
for code in list(assign):
    if code.startswith(SKIP_PREFIX):
        assign[code] = ([], [])
for code, (tc, pc) in FORCE.items():
    if code in subj_codes:
        assign[code] = ([by_code[c.upper()].id for c in tc if c.upper() in by_code],
                        [by_code[c.upper()].id for c in pc if c.upper() in by_code])

print("\n=== RESUELTOS: %d módulos ===" % len(assign))
print("=== NO ENCONTRADOS: %d ===" % len(notfound))
for x in notfound: print("  "+x)

# reclasificar polivalentes
need_gt = set()
for code,(tro,pra) in assign.items():
    for cid in tro:
        c = Classroom.browse(cid)
        if c.gela_mota=='tailerra': need_gt.add(c.code)
    for cid in pra:
        c = Classroom.browse(cid)
        if c.gela_mota=='gela': need_gt.add(c.code)
print("\n=== RECLASIFICAR a gela_tailerra: %s ===" % ", ".join(sorted(need_gt)))

if APPLY:
    for code in sorted(need_gt):
        by_code[code.upper()].gela_mota='gela_tailerra'
    for code,(tro,pra) in assign.items():
        Subject.search([('code','=',code)],limit=1).write({
            'gela_teoria_ids':[(6,0,tro)], 'tailerra_ids':[(6,0,pra)]})
    # copias HE_ (codocencia) y DESDO_ (desdoble, por ahora misma aula que el
    # origen con 2 profes): heredan aulas del módulo origen.
    ncopy = 0
    for pref in ('HE_', 'DESDO_'):
        for h in Subject.search([('code','=like', pref.replace('_','\\_') + '%')]):
            o = Subject.search([('code','=', h.code[len(pref):])], limit=1)
            if o:
                h.write({'gela_teoria_ids':[(6,0,o.gela_teoria_ids.ids)],
                         'tailerra_ids':[(6,0,o.tailerra_ids.ids)]})
                ncopy += 1
    env.cr.commit()
    print("\n>>> APLICADO (%d módulos + %d copias HE_/DESDO_)." % (len(assign), ncopy))
else:
    print("\n>>> DRY-RUN. Nada escrito.")
