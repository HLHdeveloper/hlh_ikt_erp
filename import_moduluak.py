#!/usr/bin/env python3
"""
Importación de Moduluak desde Excel → op_subject (Odoo)
Fichero: HLH_FITXATEGIAK/AKONIMOAK BUKATU GABE.xlsx

Columnas por pestaña (ciclo):
  A: Código oficial del módulo
  B: Acrónimo/nombre corto (code en Odoo)
  C: Nombre completo (name en Odoo)
"""

import openpyxl
import psycopg2
import psycopg2.extras
from datetime import datetime

XLSX_PATH = '/tmp/moduluak.xlsx'

PG_CFG = dict(
    host='postgres19', port=5432,
    user='odoo', password='odoo123',
    dbname='kudeaketa',
)

ADMIN_UID  = 2
NAME_MAXLEN = 128


def now_ts():
    return datetime.now()


def log(msg):
    print(f"  → {msg}")


def section(title):
    print(f"\n{'='*60}\n  {title}\n{'='*60}")


def find_header_row(ws):
    """Devuelve el índice de la fila de cabecera (la que tiene 'Código' en col A)."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if row[0] == 'Código':
            return i
    return None


def clean_code(val):
    if val is None:
        return None
    s = str(val).strip()
    # Eliminar decimales de códigos numéricos (3016.0 → 3016)
    if s.endswith('.0'):
        s = s[:-2]
    return s if s else None


def clean_name(val):
    if val is None:
        return None
    s = str(val).strip()
    if len(s) > NAME_MAXLEN:
        s = s[:NAME_MAXLEN - 1] + '…'
    return s if s else None


def main():
    print(f"\nAbriendo: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH)

    print("Conectando a PostgreSQL...")
    pg = psycopg2.connect(**PG_CFG)
    pc = pg.cursor(cursor_factory=psycopg2.extras.DictCursor)

    total_created = total_skipped = total_errors = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        section(f"Ciclo: {sheet_name}")

        header_row = find_header_row(ws)
        if header_row is None:
            log(f"AVISO: no se encontró cabecera en pestaña '{sheet_name}', saltando")
            continue

        created = skipped = errors = 0

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            # Columnas: A=código oficial, B=acrónimo, C=nombre completo
            cod_oficial = clean_code(row[0])
            acronimo    = clean_code(row[1])
            nombre      = clean_name(row[2])

            # Saltar filas vacías o incompletas
            if not acronimo or not nombre:
                continue
            if acronimo.upper() in ('MODULUA', 'CÓDIGO', 'CODE'):
                continue

            # Comprobar si ya existe por código (acrónimo)
            pc.execute("SELECT id FROM op_subject WHERE code = %s", (acronimo,))
            if pc.fetchone():
                skipped += 1
                continue

            try:
                pc.execute("""
                    INSERT INTO op_subject
                        (name, code, type, subject_type, active,
                         create_uid, write_uid, create_date, write_date)
                    VALUES (%s, %s, 'theory', 'compulsory', true, %s, %s, %s, %s)
                """, (nombre, acronimo, ADMIN_UID, ADMIN_UID, now_ts(), now_ts()))
                log(f"[{sheet_name}] {acronimo} — {nombre[:60]}{'…' if len(nombre)>60 else ''}")
                created += 1
            except Exception as e:
                log(f"ERROR [{sheet_name}] {acronimo}: {e}")
                pg.rollback()
                errors += 1
                continue

        pg.commit()
        log(f"Creados: {created} | Saltados: {skipped} | Errores: {errors}")
        total_created += created
        total_skipped += skipped
        total_errors  += errors

    pg.close()

    print(f"\n{'='*60}")
    print(f"  IMPORTACIÓN COMPLETADA")
    print(f"  Total creados:  {total_created}")
    print(f"  Total saltados: {total_skipped}")
    print(f"  Total errores:  {total_errors}")
    print(f"{'='*60}\n")


if __name__ == '__main__':
    main()
